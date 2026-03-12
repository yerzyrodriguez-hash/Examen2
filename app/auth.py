from flask import Blueprint, render_template, redirect, url_for, request, flash
from flask_login import login_user, logout_user, login_required, current_user
from .models import User, Lector, Libro, Categoria, Prestamo
from .extensions import db, login_manager
import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file
from .inventario import solo_admin
from datetime import datetime

auth_bp = Blueprint('auth', __name__)
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@auth_bp.route('/registro', methods=['GET', 'POST'])
@login_required
@solo_admin
def registro():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        rol = request.form.get('rol')
        nombre = request.form.get('nombre')
        c_i = request.form.get('c_i')
        celular = request.form.get('celular')
        if User.query.filter_by(username=username).first():
            flash('El nombre de usuario ya está en uso.', 'danger')
            return redirect(url_for('auth.registro'))
        if Lector.query.filter_by(C_I=c_i).first():
            flash('Esta Cédula de Identidad ya está registrada.', 'danger')
            return redirect(url_for('auth.registro'))
        nuevo_usuario = User(username=username, role=rol)
        nuevo_usuario.set_password(password)
        db.session.add(nuevo_usuario)
        db.session.flush() 
        nuevo_lector = Lector(nombre=nombre, C_I=c_i, celular=celular, usuario_id=nuevo_usuario.id)
        db.session.add(nuevo_lector)
        db.session.commit()
        flash(f'Lector "{nombre}" registrado exitosamente en el sistema.', 'success')
        return redirect(url_for('auth.lista_usuarios'))
    return render_template('registro.html')

@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        usuario = User.query.filter_by(username=username).first()
        if usuario and usuario.check_password(password):
            login_user(usuario)
            return redirect(url_for('auth.lista_usuarios'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')

@auth_bp.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('auth.lista_usuarios'))

@auth_bp.route('/usuarios')
@login_required
def lista_usuarios():
    if current_user.role != 'admin':
        return redirect(url_for('inventario.catalogo'))
    busqueda = request.args.get('busqueda', '')
    if busqueda:
        usuarios = User.query.join(Lector).filter(
            (Lector.C_I.like(f'%{busqueda}%')) |
            (Lector.nombre.like(f'%{busqueda}%')) |
            (User.username.like(f'%{busqueda}%'))
        ).all()
    else:
        usuarios = User.query.all()
    return render_template('lista_usuarios.html', usuarios=usuarios, busqueda=busqueda)

@auth_bp.route('/exportar_usuarios')
@login_required
def exportar_usuarios():
    usuarios = User.query.all()
    datos = []
    for u in usuarios:
        datos.append({
            'C.I.': u.perfil.C_I,
            'Nombre Completo': u.perfil.nombre,
            'Nombre de Usuario': u.username,
            'Celular': u.perfil.celular or 'Sin registro',
            'Rol en Sistema': 'Administrador' if u.role == 'admin' else 'Lector'
        })
    df = pd.DataFrame(datos)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Lista de Usuarios')
        worksheet = writer.sheets['Lista de Usuarios']
        for col in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 2
    output.seek(0)
    return send_file(
        output,
        download_name="Reporte_Usuarios_Biblioteca.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
@auth_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_usuario(id):
    usuario = User.query.get_or_404(id)
    if request.method == 'POST':
        usuario.username = request.form.get('username')
        usuario.role = request.form.get('rol')
        usuario.perfil.nombre = request.form.get('nombre')
        usuario.perfil.C_I = request.form.get('c_i')
        usuario.perfil.celular = request.form.get('celular')
        nueva_password = request.form.get('password')
        if nueva_password:
            usuario.set_password(nueva_password)
        db.session.commit()
        flash(f'Datos del usuario "{usuario.username}" actualizados correctamente.', 'success')
        return redirect(url_for('auth.lista_usuarios'))
    return render_template('editar_usuario.html', usuario=usuario)

@auth_bp.route('/eliminar/<int:id>')
@login_required
def eliminar_usuario(id):
    usuario = User.query.get_or_404(id)
    if usuario.username == 'admin':
        flash('Por seguridad, no puedes eliminar al administrador principal.', 'danger')
        return redirect(url_for('auth.lista_usuarios'))
    db.session.delete(usuario.perfil)
    db.session.delete(usuario)
    db.session.commit()
    flash('El usuario y su perfil han sido eliminados del sistema.', 'success')
    return redirect(url_for('auth.lista_usuarios'))

# ============= RUTAS PARA PRÉSTAMOS =============
@auth_bp.route('/prestamos')
@login_required
@solo_admin
def lista_prestamos():
    estado = request.args.get('estado')
    
    if estado:
        prestamos = Prestamo.query.filter_by(estado=estado).all()
    else:
        prestamos = Prestamo.query.all()
    
    return render_template('prestamos.html', prestamos=prestamos)

@auth_bp.route('/prestamo/nuevo', methods=['GET', 'POST'])
@login_required
@solo_admin
def nuevo_prestamo():
    if request.method == 'POST':
        usuario_id = request.form.get('usuario_id')
        libro_id = request.form.get('libro_id')

        if not usuario_id or not libro_id:
            flash('Debe seleccionar un usuario y un libro', 'danger')
            return redirect(url_for('auth.nuevo_prestamo'))

        libro = Libro.query.get(libro_id)

        prestamos_pendientes = Prestamo.query.filter_by(
            libro_id=libro_id, 
            estado='Pendiente'
        ).count()
        
        if prestamos_pendientes >= libro.stock:
            flash(f'No hay ejemplares disponibles de "{libro.titulo}"', 'danger')
            return redirect(url_for('auth.nuevo_prestamo'))
        
        prestamo = Prestamo(
            usuario_id=usuario_id,
            libro_id=libro_id
        )
        
        try:
            db.session.add(prestamo)
            db.session.commit()
            flash(f'Préstamo registrado exitosamente', 'success')
            return redirect(url_for('auth.lista_prestamos'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al registrar préstamo: {str(e)}', 'danger')
            return redirect(url_for('auth.nuevo_prestamo'))
    
    usuarios = User.query.all()
    libros = Libro.query.all()
    return render_template('nuevo_prestamo.html', usuarios=usuarios, libros=libros)

@auth_bp.route('/prestamo/devolver/<int:id>')
@login_required
@solo_admin
def devolver_prestamo(id):
    prestamo = Prestamo.query.get_or_404(id)
    
    if prestamo.estado == 'Pendiente':
        prestamo.estado = 'Devuelto'
        prestamo.fecha_devolucion = datetime.now()
        
        try:
            db.session.commit()
            flash(f'Libro "{prestamo.libro.titulo}" devuelto exitosamente', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error al procesar devolución: {str(e)}', 'danger')
    else:
        flash('Este préstamo ya fue devuelto anteriormente', 'warning')
    
    return redirect(url_for('auth.lista_prestamos'))

# ============= RUTAS PARA REPORTES DE PRÉSTAMOS =============
@auth_bp.route('/reportes/prestamos')
@login_required
@solo_admin
def reportes_prestamos():
    return render_template('reportes_prestamos.html')

@auth_bp.route('/reportes/prestamos/exportar', methods=['POST'])
@login_required
@solo_admin
def exportar_reporte_prestamos():
    fecha_inicio = request.form.get('fecha_inicio')
    fecha_fin = request.form.get('fecha_fin')
    estado = request.form.get('estado')
    tipo_reporte = request.form.get('tipo_reporte', 'completo')

    query = Prestamo.query

    if fecha_inicio:
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        query = query.filter(Prestamo.fecha_prestamo >= fecha_inicio_dt)
    
    if fecha_fin:
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
        fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
        query = query.filter(Prestamo.fecha_prestamo <= fecha_fin_dt)

    if estado and estado != 'todos':
        query = query.filter_by(estado=estado)

    prestamos = query.all()

    if tipo_reporte == 'completo':
        return generar_reporte_completo(prestamos, fecha_inicio, fecha_fin, estado)
    elif tipo_reporte == 'resumen':
        return generar_reporte_resumen(prestamos, fecha_inicio, fecha_fin)
    elif tipo_reporte == 'vencidos':
        return generar_reporte_vencidos(prestamos)
    else:
        return redirect(url_for('auth.reportes_prestamos'))

def generar_reporte_completo(prestamos, fecha_inicio, fecha_fin, estado):
    """Genera reporte detallado de todos los préstamos"""

    datos = []
    for p in prestamos:
        datos.append({
            'ID Préstamo': p.id,
            'Fecha Préstamo': p.fecha_prestamo.strftime('%d/%m/%Y %H:%M'),
            'Fecha Devolución': p.fecha_devolucion.strftime('%d/%m/%Y %H:%M') if p.fecha_devolucion else 'Pendiente',
            'Estado': p.estado,
            'Usuario': p.usuario.perfil.nombre,
            'C.I. Usuario': p.usuario.perfil.C_I,
            'Celular': p.usuario.perfil.celular or 'Sin registro',
            'Libro': p.libro.titulo,
            'Autor': p.libro.autor,
            'Categoría': p.libro.categoria.nombre if p.libro.categoria else 'Sin categoría',
        })
    
    df = pd.DataFrame(datos)
    
    total_prestamos = len(prestamos)
    prestamos_pendientes = sum(1 for p in prestamos if p.estado == 'Pendiente')
    prestamos_devueltos = sum(1 for p in prestamos if p.estado == 'Devuelto')

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        df.to_excel(writer, index=False, sheet_name='Préstamos Detallados', startrow=4)

        worksheet = writer.sheets['Préstamos Detallados']

        worksheet.merge_cells('A1:J1')
        worksheet['A1'] = 'REPORTE PRÉSTAMOS BIBLIOTECA UAB'
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.merge_cells('A2:J2')
        worksheet['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet['A2'].font = Font(size=12, italic=True)
        worksheet['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet.merge_cells('A3:J3')
        periodo = f"Período: {fecha_inicio or 'Inicio'} a {fecha_fin or 'Actualidad'} - Estado: {estado if estado and estado != 'todos' else 'Todos'}"
        worksheet['A3'] = periodo
        worksheet['A3'].font = Font(size=11)
        worksheet['A3'].alignment = Alignment(horizontal='center', vertical='center')

        column_widths = {
            'A': 12,  # ID Préstamo
            'B': 20,  # Fecha Préstamo
            'C': 20,  # Fecha Devolución
            'D': 15,  # Estado
            'E': 25,  # Usuario
            'F': 15,  # C.I. Usuario
            'G': 15,  # Celular
            'H': 40,  # Libro
            'I': 25,  # Autor
            'J': 20,  # Categoría
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        resumen_data = {
            'Métrica': ['Total Préstamos', 'Préstamos Pendientes', 'Préstamos Devueltos', 
                       'Fecha Inicio', 'Fecha Fin', 'Filtro Estado'],
            'Valor': [total_prestamos, prestamos_pendientes, prestamos_devueltos,
                     fecha_inicio if fecha_inicio else 'Todos', 
                     fecha_fin if fecha_fin else 'Todos',
                     estado if estado and estado != 'todos' else 'Todos']
        }
        df_resumen = pd.DataFrame(resumen_data)
        df_resumen.to_excel(writer, index=False, sheet_name='Resumen', startrow=4)
        
        worksheet_resumen = writer.sheets['Resumen']
        
        worksheet_resumen.merge_cells('A1:B1')
        worksheet_resumen['A1'] = 'RESUMEN - REPORTE PRÉSTAMOS BIBLIOTECA UAB'
        worksheet_resumen['A1'].font = Font(size=14, bold=True)
        worksheet_resumen['A1'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet_resumen.merge_cells('A2:B2')
        worksheet_resumen['A2'] = f'Fecha de generación: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        worksheet_resumen['A2'].font = Font(size=11, italic=True)
        worksheet_resumen['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet_resumen.merge_cells('A3:B3')
        worksheet_resumen['A3'] = periodo
        worksheet_resumen['A3'].alignment = Alignment(horizontal='center', vertical='center')

        worksheet_resumen.column_dimensions['A'].width = 30
        worksheet_resumen.column_dimensions['B'].width = 25
    
    output.seek(0)
    
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f"Reporte_Prestamos_UAB_{fecha_actual}.xlsx"
    
    return send_file(
        output,
        download_name=nombre_archivo,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def generar_reporte_resumen(prestamos, fecha_inicio, fecha_fin):
    """Genera un reporte resumen con estadísticas"""

    total_prestamos = len(prestamos)
    prestamos_pendientes = sum(1 for p in prestamos if p.estado == 'Pendiente')
    prestamos_devueltos = sum(1 for p in prestamos if p.estado == 'Devuelto')

    prestamos_por_usuario = {}
    for p in prestamos:
        nombre_usuario = p.usuario.perfil.nombre
        if nombre_usuario in prestamos_por_usuario:
            prestamos_por_usuario[nombre_usuario] += 1
        else:
            prestamos_por_usuario[nombre_usuario] = 1
    
    prestamos_por_categoria = {}
    for p in prestamos:
        categoria = p.libro.categoria.nombre if p.libro.categoria else 'Sin categoría'
        if categoria in prestamos_por_categoria:
            prestamos_por_categoria[categoria] += 1
        else:
            prestamos_por_categoria[categoria] = 1
    
    datos_resumen = [
        ['Total Préstamos', total_prestamos],
        ['Préstamos Pendientes', prestamos_pendientes],
        ['Préstamos Devueltos', prestamos_devueltos],
        ['Porcentaje Devolución', f"{(prestamos_devueltos/total_prestamos*100):.1f}%" if total_prestamos > 0 else "0%"],
        ['Período', f"{fecha_inicio or 'Inicio'} a {fecha_fin or 'Actualidad'}"]
    ]
    
    df_resumen = pd.DataFrame(datos_resumen, columns=['Métrica', 'Valor'])
    
    df_usuarios = pd.DataFrame(
        sorted(prestamos_por_usuario.items(), key=lambda x: x[1], reverse=True),
        columns=['Usuario', 'Cantidad Préstamos']
    )
    
    df_categorias = pd.DataFrame(
        sorted(prestamos_por_categoria.items(), key=lambda x: x[1], reverse=True),
        columns=['Categoría', 'Cantidad Préstamos']
    )
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, index=False, sheet_name='Resumen General')
        df_usuarios.to_excel(writer, index=False, sheet_name='Préstamos por Usuario')
        df_categorias.to_excel(writer, index=False, sheet_name='Préstamos por Categoría')
    
    output.seek(0)
    
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(
        output,
        download_name=f"Resumen_Prestamos_{fecha_actual}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

def generar_reporte_vencidos(prestamos):
    
    fecha_actual = datetime.now()
    prestamos_vencidos = []
    
    for p in prestamos:
        if p.estado == 'Pendiente':
            dias_prestamo = (fecha_actual - p.fecha_prestamo).days
            if dias_prestamo > 15:  # Consideramos vencido después de 15 días
                prestamos_vencidos.append({
                    'ID Préstamo': p.id,
                    'Usuario': p.usuario.perfil.nombre,
                    'C.I.': p.usuario.perfil.C_I,
                    'Celular': p.usuario.perfil.celular or 'Sin registro',
                    'Libro': p.libro.titulo,
                    'Fecha Préstamo': p.fecha_prestamo.strftime('%d/%m/%Y'),
                    'Días de retraso': dias_prestamo - 15,
                    'Días totales': dias_prestamo
                })
    
    if prestamos_vencidos:
        df = pd.DataFrame(prestamos_vencidos)

        df = df.sort_values('Días de retraso', ascending=False)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Préstamos Vencidos')

            resumen = pd.DataFrame([
                ['Total préstamos vencidos', len(prestamos_vencidos)],
                ['Fecha de generación', datetime.now().strftime('%d/%m/%Y %H:%M')]
            ], columns=['Descripción', 'Valor'])
            resumen.to_excel(writer, index=False, sheet_name='Resumen')
        
        output.seek(0)
        return send_file(
            output,
            download_name=f"Prestamos_Vencidos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        flash('No hay préstamos vencidos', 'info')
        return redirect(url_for('auth.reportes_prestamos'))
    
###########################DASHBOARD################################

@auth_bp.route('/dashboard')
@login_required
def dashboard():
    
    # Estadísticas generales
    total_libros = Libro.query.count()
    total_usuarios = User.query.count()
    prestamos_activos = Prestamo.query.filter_by(estado='Pendiente').count()
    prestamos_devueltos = Prestamo.query.filter_by(estado='Devuelto').count()
    
    # Libros por categoría (para gráfico de pastel)
    categorias = Categoria.query.all()
    categorias_nombres = [c.nombre for c in categorias]
    libros_por_categoria = [len(c.libros) for c in categorias]
    
    # Préstamos por mes (últimos 6 meses)
    from datetime import datetime, timedelta
    import calendar
    
    meses = []
    prestamos_por_mes = []
    
    for i in range(5, -1, -1):
        fecha = datetime.now() - timedelta(days=30*i)
        mes_nombre = calendar.month_name[fecha.month][:3] + f" {fecha.year}"
        meses.append(mes_nombre)
        
        inicio_mes = datetime(fecha.year, fecha.month, 1)
        if fecha.month == 12:
            fin_mes = datetime(fecha.year + 1, 1, 1) - timedelta(days=1)
        else:
            fin_mes = datetime(fecha.year, fecha.month + 1, 1) - timedelta(days=1)
        
        cantidad = Prestamo.query.filter(
            Prestamo.fecha_prestamo >= inicio_mes,
            Prestamo.fecha_prestamo <= fin_mes
        ).count()
        prestamos_por_mes.append(cantidad)
    
    # Top 5 libros más prestados
    from sqlalchemy import func
    top_libros = db.session.query(
        Libro.titulo,
        func.count(Prestamo.id).label('total_prestamos')
    ).join(Prestamo).group_by(Libro.id).order_by(func.count(Prestamo.id).desc()).limit(5).all()
    
    top_libros_nombres = [l[0] for l in top_libros]
    top_libros_cantidad = [l[1] for l in top_libros]
    
    # Usuarios con más préstamos
    top_usuarios = db.session.query(
        Lector.nombre,
        func.count(Prestamo.id).label('total_prestamos')
    ).join(User, Lector.usuario_id == User.id).join(Prestamo, User.id == Prestamo.usuario_id).group_by(Lector.id).order_by(func.count(Prestamo.id).desc()).limit(5).all()
    
    top_usuarios_nombres = [u[0] for u in top_usuarios]
    top_usuarios_cantidad = [u[1] for u in top_usuarios]
    
    return render_template(
        'dashboard.html',
        total_libros=total_libros,
        total_usuarios=total_usuarios,
        prestamos_activos=prestamos_activos,
        prestamos_devueltos=prestamos_devueltos,
        categorias_nombres=categorias_nombres,
        libros_por_categoria=libros_por_categoria,
        meses=meses,
        prestamos_por_mes=prestamos_por_mes,
        top_libros_nombres=top_libros_nombres,
        top_libros_cantidad=top_libros_cantidad,
        top_usuarios_nombres=top_usuarios_nombres,
        top_usuarios_cantidad=top_usuarios_cantidad
    )